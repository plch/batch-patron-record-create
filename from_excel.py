from openpyxl import load_workbook
from configparser import ConfigParser
from base64 import b64encode
import requests
import json

# get the api paramaters from the configuration file
config = ConfigParser()
config.read('api_info.ini')
base_url = config['api']['base_url']
client_key = config['api']['client_key']
client_secret = config['api']['client_secret']

class Patron:
    def __init__(self,
        expiration_date,
        patron_codes_p1,
        patron_codes_p2,
        patron_codes_p3,
        patron_codes_p4,
        note,
        patron_type,
        birth_date,
        home_library_code,
        block_info,
        last_name,
        first_name,
        addresses_line1,
        addresses_line2,
        phone,
        barcode,
        patron_message,
        pin
    ):
        # patron data will be a dictionary that will be converted to json and passed to the API
        self.patron_data = {}
        self.patron_data = {
            # "expirationDate": str(expiration_date),
            "expirationDate": expiration_date.strftime('%Y-%m-%d'),
            "patronCodes": {
 		        "pcode1": str(patron_codes_p1),
                "pcode2": str(patron_codes_p2),
                "pcode3": int(patron_codes_p3),
                "pcode4": int(patron_codes_p4)
            },
            "varFields": [{
     		     "fieldTag": "x",
     		     "content": str(note)
            }],
            "patronType": int(patron_type),
            # "birthDate": str(birth_date),
            "birthDate": birth_date.strftime('%Y-%m-%d'),
            "homeLibraryCode": str(home_library_code),
            "blockInfo": {
 		         "code": str(block_info)
            },
            "names": [
 		         str(last_name) + ", " + str(first_name)
            ],
            "addresses": [{
                "lines": [
                    str(addresses_line1),
                    str(addresses_line2)
                ],
                "type": "a"
            }],
            "phones": [{
                "number": str(phone),
                "type": "t"
            }],
            "barcodes": [
                str(barcode)
            ],
            "fixedFields": {
                "54": {
                    "label": "Patron Message",
                    "value": str(patron_message)
                }
            },
            "pin": str(pin),
        }

    def get_dict(self):
        return self.patron_data

wb = load_workbook(filename = 'Columbia Tusculum Card Data.xlsx')

print(wb.sheetnames[0])
ws = wb.active

print('sheet dimensions: ')
print('rows:' + str(ws.max_row) )
print('columns: ' + str(ws.max_column) )


# we need to base64 encode our key and secret for the basic auth done for getting the token
# base64 expects the string to be in ascii format, and we have to decode it back to utf-8 for the header
auth_string = b64encode(
    (client_key + ':' + client_secret).encode('ascii')
).decode('utf-8')

# create dictionary for the headers, and add our auth_string to it
headers = {}
headers['authorization'] = 'basic ' + auth_string

# get the response from the token endpoint
r = requests.request('POST', base_url + '/token', headers=headers)

# convert the json response into a python object
json_data = json.loads(r.text)
# note: we could also use the following .json() method built into requests .. but it may not work for older versions
# json_data = r.json()
# print('token: ' + json_data['access_token'])

# reset the headers and prepare it with the bearer token we received from the authorization request
headers = {}
headers['authorization'] = 'bearer ' + json_data['access_token']
headers['content-type'] = 'application/json'
headers['accept'] = 'application/json'

print('headers: ')
print(headers)

# send the next request for token info with our bearer token
r = requests.request("GET", base_url + '/info/token', headers=headers)
print(r.text)



# for row in range(1, ws.max_row + 1):
#     print(ws['b' + str(row)].value)

# working
# for row in ws.rows:
#     print()
#     print("--")
#     print()
#     for cell in row:
#         print(cell.value)

# itterrate over the row (note, for now, i'm deleting the first row from the spreadsheet ... the offset_row=1 appears to jump past the last row)
for row in ws.iter_rows():
    # convert the values of each cell into a list of values
    row_list = list(cell.value for cell in row)
    # print(row_list)

    # create the patron dictionary, unpacking the list as arguments for the Patron class
    patron = Patron(*row_list)
    patron_data = patron.get_dict()
    print(json.dumps(patron_data), '\n')

    r = requests.request('POST', base_url + '/patrons', headers=headers, data=json.dumps(patron_data))
    print(r.status_code)
    print(r.text)


# for cell in ws[2]:
#     fields[count] = cell.value
#     count += 1
#     print(cell.value)
